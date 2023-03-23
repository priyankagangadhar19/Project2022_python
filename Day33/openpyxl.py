import openpyxl

#open the excel file-workbook
wb=openpyxl.load_workbook('F:/B64/Book1.xlsx')
rc=wb['Sheet2'].max_row
print(rc)

cc= wb['Sheet2'].max_column
print(cc)
for i in range(1,rc+1):
    for j in range(1,cc+1):
        v=wb['Sheet2'].cell(i,j).value
        print(v,end=' ')
    print()

wb.close()










